# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'H:\Developers\python projects\Selenium_gitam_marks\fetch.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(645, 364)
        img_path=str(Path(__file__).parent.absolute())+"\\student.png"
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(img_path), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainWindow.setWindowIcon(icon)
        MainWindow.setAutoFillBackground(False)
        MainWindow.setStyleSheet("background-color: rgb(0, 0, 0);\n""font: 20pt \"Bebas Neue\";\n""color: rgb(255, 255, 255);")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(10, 20, 151, 41))
        self.label_2.setStyleSheet("font: 20pt \"Bebas Neue\";\n""color: rgb(255, 255, 255);")
        self.label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(10, 80, 151, 41))
        self.label_3.setStyleSheet("font: 20pt \"Bebas Neue\";\n""color: rgb(255, 255, 255);")
        self.label_3.setAlignment(QtCore.Qt.AlignCenter)
        self.label_3.setObjectName("label_3")
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(10, 140, 151, 41))
        self.label_4.setStyleSheet("font: 20pt \"Bebas Neue\";\n""color: rgb(255, 255, 255);")
        self.label_4.setAlignment(QtCore.Qt.AlignCenter)
        self.label_4.setObjectName("label_4")
        self.sembox = QtWidgets.QSpinBox(self.centralwidget)
        self.sembox.setGeometry(QtCore.QRect(180, 140, 91, 41))
        font = QtGui.QFont()
        font.setFamily("Bebas Neue")
        font.setPointSize(20)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.sembox.setFont(font)
        self.sembox.setStyleSheet("background-color: rgb(202, 202, 202);\n""color: rgb(0, 0, 0);")
        self.sembox.setObjectName("sembox")
        self.excelbutton = QtWidgets.QPushButton(self.centralwidget)
        self.excelbutton.setGeometry(QtCore.QRect(300, 140, 331, 41))
        self.excelbutton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.excelbutton.setStyleSheet("font: 15pt \"Arial Rounded MT Bold\";\n""color: rgb(0, 0, 0);\n""background-color: rgb(202, 202, 202);")
        self.excelbutton.setObjectName("excelbutton")
        self.progressBar = QtWidgets.QProgressBar(self.centralwidget)
        self.progressBar.setGeometry(QtCore.QRect(20, 270, 611, 23))
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")
        self.startbutton = QtWidgets.QPushButton(self.centralwidget)
        self.startbutton.setEnabled(True)
        self.startbutton.setGeometry(QtCore.QRect(20, 210, 191, 41))
        self.startbutton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.startbutton.setStyleSheet("font: 20pt \"Bebas Neue\";\n""color: rgb(0, 0, 0);\n""background-color: rgb(202, 202, 202);")
        self.startbutton.setObjectName("startbutton")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(20, 320, 611, 31))
        font = QtGui.QFont()
        font.setFamily("Bebas Neue")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label.setFont(font)
        self.label.setStyleSheet("font: 90 12pt \"Myanmar Text\";")
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        self.srollno = QtWidgets.QLineEdit(self.centralwidget)
        self.srollno.setGeometry(QtCore.QRect(180, 20, 451, 41))
        self.srollno.setContextMenuPolicy(QtCore.Qt.DefaultContextMenu)
        self.srollno.setStyleSheet("font: 20pt \"Bebas Neue\";\n""color: rgb(0, 0, 0);\n""background-color: rgb(202, 202, 202);")
        self.srollno.setClearButtonEnabled(False)
        self.srollno.setObjectName("srollno")
        self.srollno.setMaxLength(12)
        self.erollno = QtWidgets.QLineEdit(self.centralwidget)
        self.erollno.setGeometry(QtCore.QRect(180, 80, 451, 41))
        self.erollno.setStyleSheet("font: 20pt \"Bebas Neue\";\n""color: rgb(0, 0, 0);\n""background-color: rgb(202, 202, 202);")
        self.erollno.setObjectName("erollno")
        self.erollno.setMaxLength(12)
        self.error_box = QtWidgets.QLabel(self.centralwidget)
        self.error_box.setEnabled(True)
        self.error_box.setGeometry(QtCore.QRect(240, 210, 391, 41))
        self.error_box.setStyleSheet("color: rgb(255, 0, 0);\n""font: 90 14pt \"Myanmar Text\";")
        self.error_box.setText("")
        self.error_box.setAlignment(QtCore.Qt.AlignCenter)
        self.error_box.setObjectName("error_box")
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        self.startbutton.clicked.connect(self.run)
        self.excelbutton.clicked.connect(self.open_check)

    def open_check(self):
        
        fileName = QtWidgets.QFileDialog.getOpenFileName(None, 'OpenFile',"","All Files (*)")
        global pathx
        pathx=fileName[0]
        self.error_box.setText("Checking-length")
        time.sleep(1)
        self.progressBar.setProperty("value", 20)
        self.error_box.setText("Checking-SEM-Validity")
        time.sleep(1)
        self.progressBar.setProperty("value", 42)
        self.error_box.setText("Checking-path")
        time.sleep(1)
        self.progressBar.setProperty("value", 57)
        self.error_box.setText("Connecting...")
        time.sleep(1)
        self.progressBar.setProperty("value", 78)
        if((len(self.srollno.text())==12) and (len(self.erollno.text())==12) and (int(self.sembox.text())>0) and (pathx.find('xlsx')!=-1)):
            self.error_box.setStyleSheet("color: rgb(0, 255, 0);\n""font: 90 14pt \"Myanmar Text\";")
            self.error_box.setText("All-ok")
            self.progressBar.setProperty("value", 100)
            
        else:
            self.error_box.setStyleSheet("color: rgb(255, 0, 0);\n""font: 90 14pt \"Myanmar Text\";")
            self.error_box.setText("Invalid Details")
            self.progressBar.setProperty("value", 0)
        try:
            sr=int(self.srollno.text())
            er=int(self.erollno.text())
        except:
            self.error_box.setStyleSheet("color: rgb(255, 0, 0);\n""font: 90 14pt \"Myanmar Text\";")
            self.error_box.setText("Invalid Details")
            self.progressBar.setProperty("value", 0)

    def run(self):
        sr=int(self.srollno.text())
        er=int(self.erollno.text())
        self.error_box.setText("All-ok")
        roll_no= sr
        workbook=openpyxl.load_workbook(pathx) 
        sheet=workbook.active 
        r=2
        driver_path=str(Path(__file__).parent.absolute())+"\\chromedriver.exe"
        browser = webdriver.Chrome(driver_path) #chrome-driver
        browser.get('https://doeresults.gitam.edu/onlineresults/pages/Newgrdcrdinput1.aspx') #redirecting-to-university-results-portal
        while roll_no<=er:
            sem_drop_box=Select(browser.find_element_by_xpath("//*[@id='cbosem']")) #selecting-semester-drop-box
            sem_drop_box.select_by_visible_text(self.sembox.text()) #select-semester
            roll_no_box=browser.find_element_by_xpath("//*[@id='txtreg']") #selecting-roll-no-box
            roll_no_box.clear() #clearing-unwanted-text
            roll_no_box.send_keys(roll_no) #filling-roll-no
            browser.find_element_by_xpath("//*[@id='Button1']").click() #selecting-get-result-butoon-and-clicking-it
            time.sleep(1)
            if(browser.current_url == 'https://doeresults.gitam.edu/onlineresults/pages/View_Result_Grid.aspx'):

                name=browser.find_element_by_xpath("//*[@id='lblname']")
                roll=browser.find_element_by_xpath("//*[@id='lblregdno']")
                cgpa=browser.find_element_by_xpath("//*[@id='lblgpa']")
                gpa=browser.find_element_by_xpath("//*[@id='lblcgpa']")
                
                sheet.cell(row=r,column=1).value=name.text
                sheet.cell(row=r,column=2).value=roll.text
                sheet.cell(row=r, column=3).value=cgpa.text
                sheet.cell(row=r, column=4).value=gpa.text
                roll_no=roll_no+1
                r=r+1
                browser.back()
            else:
                sheet.cell(row=r, column=1).value="NOT-FOUND"
                sheet.cell(row=r, column=2).value="NOT-FOUND"
                sheet.cell(row=r, column=3).value="NOT-FOUND"
                sheet.cell(row=r, column=4).value="NOT-FOUND"
                r=r+1
                roll_no=roll_no+1
        workbook.save(pathx)
        browser.close()


        #print(self.srollno.text())
        #print(self.erollno.text())
        #print(self.sembox.text())
        #print(path)
        

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Class Marks Fetcher By-Asish"))
        self.label_2.setText(_translate("MainWindow", "Start roll-no:"))
        self.label_3.setText(_translate("MainWindow", "End roll-no:"))
        self.label_4.setText(_translate("MainWindow", "Select sem"))
        self.excelbutton.setText(_translate("MainWindow", "Select Excel File And Validate"))
        self.startbutton.setText(_translate("MainWindow", "Get Data"))
        self.label.setText(_translate("MainWindow", "Powered by Python- Coded by Asish"))


if __name__ == "__main__":
    import sys
    import openpyxl
    from selenium import webdriver
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import Select
    from pathlib import Path
    import time
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

